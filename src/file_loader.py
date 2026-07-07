# src/file_loader.py
#
# Phase 6 ingestion front-door: parse CSV/TSV/JSON/XLSX uploads into a common
# list[dict] behind one interface (load_rows), then map + reject-malformed via
# ingest_file. No value normalization here — that is Phase 8's job.
import csv
import json
from pathlib import Path
from typing import Any, Dict, List

import yaml

from src.column_mapper import map_row
from src.schemas import IngestBatch, RejectedRow

_MAPPING_PATH = "config/column_mapping.yaml"


def load_rows(path: str) -> List[Dict[str, Any]]:
    """Auto-detect format by file extension and return a list of raw dicts.

    .csv/.tsv -> csv.DictReader (utf-8-sig, BOM-safe)
    .json     -> stdlib json (top-level list OR {"contacts":[...]}/{"rows":[...]})
    .xlsx/.xls -> openpyxl (read_only, data_only; header row -> keys)
    Anything else raises ValueError naming the unsupported extension.
    """
    suffix = Path(path).suffix.lower()
    if suffix in (".csv", ".tsv"):
        return _load_csv(path, delimiter="\t" if suffix == ".tsv" else ",")
    if suffix == ".json":
        return _load_json(path)
    if suffix in (".xlsx", ".xls"):
        return _load_xlsx(path)
    raise ValueError(f"Unsupported file extension: {suffix or '(none)'}")


def _load_csv(path: str, delimiter: str) -> List[Dict[str, Any]]:
    # utf-8-sig transparently strips an Excel-exported BOM so the first header is clean.
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(row) for row in reader]


def _load_json(path: str) -> List[Dict[str, Any]]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("contacts", "rows"):
            if isinstance(data.get(key), list):
                return data[key]
    raise ValueError(
        "Unrecognized JSON shape: expected a top-level list or an object with a "
        "'contacts' or 'rows' list."
    )


def _load_xlsx(path: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        headers = ["" if cell is None else str(cell) for cell in header]
        out: List[Dict[str, Any]] = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue  # fully blank row
            # blank cell -> "" so a blank xlsx cell matches a blank csv/json field
            out.append({h: ("" if v is None else v) for h, v in zip(headers, row)})
        return out
    finally:
        wb.close()


def _has_identity(mapped: Dict[str, Any], required: Dict[str, Any]) -> bool:
    # True when, for some group in required["any_of"], every key is present and non-empty.
    for group in required.get("any_of", []):
        if all(mapped.get(key) not in (None, "") for key in group):
            return True
    return False


def ingest_file(path: str) -> IngestBatch:
    """load -> map -> split into accepted canonical rows and structured rejects.

    Per-row try/except so one bad row can never crash the batch. A row yielding
    no identity key lands in rejects with row_index + reason 'no identity key'.
    """
    with open(_MAPPING_PATH, encoding="utf-8") as f:
        mapping = yaml.safe_load(f)
    required = mapping.get("required_identity", {})

    accepted: List[Dict[str, Any]] = []
    rejects: List[RejectedRow] = []

    for i, raw in enumerate(load_rows(path)):
        try:
            if not isinstance(raw, dict):
                rejects.append(
                    RejectedRow(row_index=i, reason="row is not an object", raw={"value": repr(raw)})
                )
                continue
            mapped = map_row(raw, mapping)
            if not _has_identity(mapped, required):
                rejects.append(RejectedRow(row_index=i, reason="no identity key", raw=raw))
                continue
            accepted.append(mapped)
        except Exception as e:  # one bad row must never crash the batch
            safe_raw = raw if isinstance(raw, dict) else {"value": repr(raw)}
            rejects.append(RejectedRow(row_index=i, reason=f"parse error: {e}", raw=safe_raw))

    return IngestBatch(rows=accepted, rejects=rejects)
