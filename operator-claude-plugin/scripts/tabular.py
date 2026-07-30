"""operator-claude-plugin/scripts/tabular.py

Reads a CSV or XLSX file exactly as the operator provided it — no header cleaning, no
column mapping, no normalization. That is n8n's `Map Columns` node's job, fed the
unchanged file (D-07, STRUCT-01). This module also converts an XLSX source to CSV bytes
for the wire, since `Extract From File` only parses `operation: "csv"` — that conversion
changes format only, never column names or values.
"""
import csv
import io
import json
from pathlib import Path


class UnsupportedFileError(ValueError):
    """Raised for any file extension this plugin does not read."""


def read_table(path) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) for a .csv or .xlsx file. Headers are verbatim — no
    cleaning, no normalization, no mapping."""
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], []
        return rows[0], rows[1:]

    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                return [], []
            headers = ["" if c is None else str(c) for c in header_row]
            rows = [
                ["" if c is None else str(c) for c in r]
                for r in it
                if any(c is not None for c in r)
            ]
            return headers, rows
        finally:
            wb.close()

    raise UnsupportedFileError(f"Unsupported file extension: {suffix or '(none)'}")


def to_csv_bytes(path) -> bytes:
    """The CSV bytes to send over the wire: original bytes for a .csv source, a
    re-serialized (format-only, never remapped) CSV for an .xlsx source."""
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return path.read_bytes()

    if suffix == ".xlsx":
        headers, rows = read_table(path)
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        return buf.getvalue().encode("utf-8")

    raise UnsupportedFileError(f"Unsupported file extension: {suffix or '(none)'}")


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 2:
        print(json.dumps({"ok": False, "error": "usage: tabular.py <path>"}))
        raise SystemExit(1)
    try:
        _headers, _rows = read_table(sys.argv[1])
    except Exception as _e:
        print(json.dumps({"ok": False, "error": str(_e)}))
        raise SystemExit(1)
    print(json.dumps({"ok": True, "headers": _headers, "row_count": len(_rows)}))
